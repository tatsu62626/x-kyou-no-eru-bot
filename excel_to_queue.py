"""
Excel連携スクリプト

「今日のエール_投稿カレンダー.xlsx」の投稿カレンダーシートを読み込み、
ステータスが「下書き」かつ投稿内容が入っている行を content_queue.csv に登録する。
登録した行はExcel側のステータスを「予約済み」に更新する（重複登録防止）。

使い方:
  python excel_to_queue.py
  → 同じフォルダの「今日のエール_投稿カレンダー.xlsx」を読み込み、content_queue.csvを更新

  python excel_to_queue.py 別のファイル名.xlsx
  → ファイル名を指定して実行

その後 `python scheduler.py` を起動しておけば、予約時刻が来た投稿が自動投稿されます。
"""
import sys
import datetime
import logging

from openpyxl import load_workbook

from config import Config
from content_queue import ContentQueue, COLUMNS
import pandas as pd

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

DEFAULT_EXCEL = "今日のエール_投稿カレンダー.xlsx"
SHEET_NAME = "投稿カレンダー"

COL_DATE = 1       # A列: 日付
COL_SLOT = 3        # C列: スロット
COL_TIME = 4        # D列: 投稿時刻
COL_CONTENT = 5     # E列: 投稿内容
COL_STATUS = 6      # F列: ステータス


def existing_ids(csv_path):
    try:
        df = pd.read_csv(csv_path, dtype=str)
        return set(df["id"].fillna(""))
    except FileNotFoundError:
        return set()


def sync(excel_path):
    wb = load_workbook(excel_path)
    ws = wb[SHEET_NAME]

    queue = ContentQueue(Config.QUEUE_CSV)
    already_queued = existing_ids(Config.QUEUE_CSV)

    added = 0
    skipped_no_content = 0
    skipped_status = 0
    skipped_duplicate = 0

    for row in range(2, ws.max_row + 1):
        status = ws.cell(row=row, column=COL_STATUS).value
        content = ws.cell(row=row, column=COL_CONTENT).value
        date_val = ws.cell(row=row, column=COL_DATE).value
        time_val = ws.cell(row=row, column=COL_TIME).value

        if not content or not str(content).strip():
            skipped_no_content += 1
            continue
        if status != "下書き":
            skipped_status += 1
            continue
        if date_val is None or time_val is None:
            continue

        date_only = date_val.date() if isinstance(date_val, datetime.datetime) else date_val
        scheduled_dt = datetime.datetime.combine(date_only, time_val)
        scheduled_str = scheduled_dt.strftime("%Y-%m-%d %H:%M")
        post_id = f"{scheduled_dt.strftime('%Y%m%d')}_{time_val.strftime('%H%M')}"

        if post_id in already_queued:
            skipped_duplicate += 1
            # 既にキュー登録済みなのでExcel側だけステータスを合わせておく
            ws.cell(row=row, column=COL_STATUS, value="予約済み")
            continue

        queue.add_post(post_id, scheduled_str, str(content).strip())
        ws.cell(row=row, column=COL_STATUS, value="予約済み")
        already_queued.add(post_id)
        added += 1

    wb.save(excel_path)

    logger.info(f"新規登録: {added}件")
    logger.info(f"スキップ(内容未入力): {skipped_no_content}件")
    logger.info(f"スキップ(下書き以外のステータス): {skipped_status}件")
    logger.info(f"スキップ(登録済み・重複回避): {skipped_duplicate}件")
    logger.info(f"→ {Config.QUEUE_CSV} と {excel_path} を更新しました。")


if __name__ == "__main__":
    excel_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_EXCEL
    sync(excel_path)
